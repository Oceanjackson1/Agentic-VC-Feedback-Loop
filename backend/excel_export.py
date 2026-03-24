"""Excel 导出模块：按场景动态表头导出 .xlsx"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from prompts import get_scenario


def to_excel(items: list[dict], lang: str = "zh", scenario_id: str = "vc_pitch") -> bytes:
    """
    将分析结果 JSON 数组转为 Excel 文件，返回 bytes。
    根据 scenario_id 动态选择表头和行提取逻辑。
    """
    scenario = get_scenario(scenario_id)

    wb = Workbook()
    ws = wb.active
    ws.title = scenario.SHEET_NAME_EN if lang == "en" else scenario.SHEET_NAME_ZH

    headers = scenario.EXCEL_HEADERS_EN if lang == "en" else scenario.EXCEL_HEADERS_ZH
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx, item in enumerate(items, 2):
        row_data = scenario.extract_row(item)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
